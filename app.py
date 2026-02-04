import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import re
import io

# --- LOGIC FUNCTIONS ---
def extract_base_number(text):
Â  Â  if pd.isna(text): return None
Â  Â  match = re.search(r'(\d+)', str(text))
Â  Â  return match.group(1) if match else None

def format_val(val):
Â  Â  try:
Â  Â  Â  Â  return f"{float(val):.4f}"
Â  Â  except:
Â  Â  Â  Â  return str(val)

# --- THE WEBSITE INTERFACE ---
st.set_page_config(page_title="CMM to IR Converter", page_icon="ðŸ“Š")

st.title("ðŸ“Š CMM Result to IR Automator")
st.write("Hello! The app is officially running. Upload your files below.")

# 1. FILE UPLOADERS
uploaded_cmm = st.file_uploader("Step 1: Upload CMM Result (Excel)", type=["xlsx"])
uploaded_template = st.file_uploader("Step 2: Upload IR Template (Excel)", type=["xlsx"])

if uploaded_cmm and uploaded_template:
Â  Â  if st.button("ðŸš€ Process and Generate Report"):
Â  Â  Â  Â  with st.spinner("Processing data..."):
Â  Â  Â  Â  Â  Â  try:
Â  Â  Â  Â  Â  Â  Â  Â  # --- READ CMM DATA ---
Â  Â  Â  Â  Â  Â  Â  Â  df_raw = pd.read_excel(uploaded_cmm, header=None, nrows=50)
Â  Â  Â  Â  Â  Â  Â  Â  header_row_idx = next((i for i, row in df_raw.iterrows() if row.astype(str).str.contains("Characteristic", case=False).any()), None)
Â  Â  Â  Â  Â  Â  Â  Â Â 
Â  Â  Â  Â  Â  Â  Â  Â  if header_row_idx is None:
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  st.error("Could not find 'Characteristic' column in CMM file.")
Â  Â  Â  Â  Â  Â  Â  Â  else:
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  df_cmm = pd.read_excel(uploaded_cmm, header=header_row_idx)
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  df_cmm.columns = [str(c).strip().upper() for c in df_cmm.columns]
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â Â 
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  cmm_results = {}
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  for _, row in df_cmm.iterrows():
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  raw_text = str(row.get("CHARACTERISTIC", "")).strip().upper()
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  base_num = extract_base_number(raw_text)
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  if not base_num: continue
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â Â 
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  is_coordinate = any(raw_text.endswith(f".{c}") or raw_text.endswith(f" {c}") or raw_text == c for c in ['X', 'Y', 'Z'])
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  if is_coordinate: continue

Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  try:
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  val = float(row.get("ACTUAL", 0))
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  except:
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  continue

Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  if base_num not in cmm_results:
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  cmm_results[base_num] = {'master': None, 'samples': []}

Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  if raw_text == base_num or raw_text == f"{base_num}.0":
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  cmm_results[base_num]['master'] = val
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  else:
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  cmm_results[base_num]['samples'].append(val)

Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  # --- FILL TEMPLATE ---
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  wb = load_workbook(uploaded_template)
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  ws = wb.active
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â Â 
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  id_col_idx, res_col_idx, start_row = None, None, None
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  for r in range(1, 30):
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  for c in range(1, ws.max_column + 1):
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  cell_val = str(ws.cell(row=r, column=c).value)
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  if "5. Char No." in cell_val: id_col_idx, start_row = c, r
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  if "9. Results" in cell_val: res_col_idx = c

Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  for r in range(start_row + 1, ws.max_row + 1):
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  ir_id = extract_base_number(ws.cell(row=r, column=id_col_idx).value)
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  if ir_id in cmm_results:
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  data = cmm_results[ir_id]
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  if data['master'] is not None:
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  final_output = format_val(data['master'])
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  elif data['samples']:
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  vals = data['samples']
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  final_output = f"{format_val(min(vals))} - {format_val(max(vals))}" if len(vals) > 1 and min(vals) != max(vals) else format_val(vals[0])
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  else: continue
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  ws.cell(row=r, column=res_col_idx).value = final_output

Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  # SAVE TO MEMORY FOR DOWNLOAD
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  output = io.BytesIO()
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  wb.save(output)
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  output.seek(0)
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â Â 
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  st.success("âœ… Report Generated Successfully!")
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  st.download_button(
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  label="ðŸ“¥ Download Final IR Report",
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  data=output,
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  file_name="Final_Report_Done.xlsx",
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  )
Â  Â  Â  Â  Â  Â  except Exception as e:
Â  Â  Â  Â  Â  Â  Â  Â  st.error(f"An error occurred: {e}")
